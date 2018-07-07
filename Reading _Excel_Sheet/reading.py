#### This import section is where my mistake was at
#### This works for me
import openpyxl    ### Excel files 
from openpyxl.styles import PatternFill, Border, Side, Alignment, Protection, Font
from openpyxl.styles import Fill, Color
# from openpyxl.styles import Style
from openpyxl.styles.colors import RED
from openpyxl.styles.colors import GREEN


str_xls_PathFileCurrent = str_xls_FileName
### Opens Excel Document
var_xls_FileOpen    = openpyxl.load_workbook(str_xls_PathFileCurrent) 
### Opens up the Excel worksheet 
var_xls_TabName     = var_xls_FileOpen.worksheets[0]                  
### Put the spreadsheet tab names into an array 
ary_xls_SheetNames  = var_xls_FileOpen.get_sheet_names()              
### Open the sheet in the file you working on 
var_xls_TabSheet    = var_xls_FileOpen.get_sheet_by_name(ary_xls_SheetNames[0])
xls_cell = var_xls_TabSheet['d10']

#### Changes the cell background color 
xls_cell.style = Style(fill=PatternFill(patternType='solid'
    , fgColor=Color('C4C4C4')))  ### Changes background color 

#### Changes the fonts (does not use style) 
xls_cell.font = xls_cell.font.copy(color  = 'FFFF0000') ### Works (Changes to red font text) 
xls_cell.font = xls_cell.font.copy(bold  = True) ### Works (Changes to bold font) 
xls_cell.font = xls_cell.font.copy(italic= True) ### Works (Changes to Italic Text) 
xls_cell.font = xls_cell.font.copy(size  =   34) ### Works (Changes Size) 