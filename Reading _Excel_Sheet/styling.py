from openpyxl import Workbook
wb = Workbook("Chandoor_Input.xlsx")

# grab the active worksheet
ws = wb.active
# ws[1, 'f'] = color.Red
# Data can be assigned directly to cells
# ws['A10'] = 42

# Rows can also be appended
# ws.append([1, 2, 3])

# Python types will automatically be converted
# import datetime
# ws['A2'] = datetime.datetime.now()

# Save the file
wb.save("sample.xlsx")